from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

#To manage patients from admin mode.
def heading(): #Give the headings in first row.
    workbook = load_workbook(r"C:\Users\maashree\Desktop\Admin\patient_database.xlsx")
    worksheet = workbook.active
    
    worksheet["A1"].value = "Patient ID"
    worksheet["B1"].value = "Name"
    worksheet["C1"].value = "Gender"
    worksheet["D1"].value = "Room No"
    worksheet["E1"].value = "Address"
    worksheet["F1"].value = "Doctor's Name"
    worksheet["G1"].value = "Department"
    worksheet["H1"].value = "Fee to be Deposited"
    workbook.save(r"C:\Users\maashree\Desktop\Admin\patient_database.xlsx")

def add():
    
        workbook = load_workbook(r"C:\Users\maashree\Desktop\Admin\patient_database.xlsx") #Initialise the database in a variable
        worksheet = workbook.active #Initialise the current worksheet in a variable.

        try: #To avoid non integer input.
            number = int(input("How many records do you want to add?: "))
        except:
            print("No. of records must be an integer value.")
            number = int(input("How many records do you want to add?: "))
            


        print("Enter the details below:")
        for i in range(number):
            check = lastID() #Return value of largest/last ID in check
            if check == None: #To make first entry of database have ID as 1, because value of lastID=None means database empty.
                ID = 1
            else: # To make ID as 1+row number because 1st row occupied by heading, if database contains atleast one (or more) values.
                ID = check+1
                
            
            name = input("Enter name : ")
            gender = input("Enter gender : ")
            try: #To avoid non integer input
                room_no = int(input("Enter room number : "))
            except:
                print("Room number should be an integer : ")
                room_no = int(input("Enter room number : "))

            address = input("Enter address : ")
            doctor = input("Enter the name of the doctor handling the case : ")
            department = input("Enter department : ")
            try:
                fee = float(input("Enter the amount of fee that is to be deposited by patient (in $) : "))
            except:
                print("Fee amount should be an integer or a floating point number.")
                fee = float(input("Enter the amount of fee that is to be deposited by patient (in $) : "))
            data = [ID, name, gender, room_no, address, doctor, department, "$"+str(fee)]
            worksheet.append(data)
            workbook.save(r"C:\Users\maashree\Desktop\Admin\patient_database.xlsx")
            print("The record of patient", ID, "is added")
            
def lastID():
    workbook = load_workbook(r"C:\Users\maashree\Desktop\Admin\patient_database.xlsx")
    worksheet = workbook.active
    L = []
    for i in range(2, 10000): #Forms a list of all values of column A if the values are digits. 
        ID = worksheet["A"+str(i)].value
        IDs = str(ID)
        if IDs.isdigit():
            L.append(ID)
            continue
        
    try: #The maximum value is the last ID.
        last = max(L)
        print("Last ID is : ", last)
        return last
    except: #If list is filled with None, database is empty.
        print("Database is currently empty, last ID is 0")


def display():
    workbook = load_workbook(r"C:\Users\maashree\Desktop\Admin\patient_database.xlsx")
    worksheet = workbook.active

    try: #To avoid non integer input.
        ID = int(input("Enter the patient ID whose records you want to see: "))
    except:
        print("The ID of patient must be an integer.")
        ID = int(input("Enter the patient ID whose records you want to see: "))
        
    start = ID+1 #ID 1 will be on row 2 because 1st row occupied by headings
    stop = ID+2 #One more than row number to pick one value by for loop
    header = 1 #Row number of headings.
    for row in range(start,stop): #runs only once.
        for col in range(1,9): 
            char = get_column_letter(col) #Gives letter corresponding to no. like 1 gives A, 2 gives B etc. 
            print(worksheet[char+str(header)].value," : ",worksheet[char + str(row)].value)

def edit():
    workbook = load_workbook(r"C:\Users\maashree\Desktop\Admin\patient_database.xlsx")
    worksheet = workbook.active

    try: #Avoid non integer input
        ID = int(input("Enter the patient ID whose records you want to edit: "))
    except:
        print("The ID of patient must be an integer.")
        ID = int(input("Enter the patient ID whose records you want to edit: "))
    

    start = ID+1 #To pick one row in for loop.
    stop = ID+2
    
    print("Enter the details to be edited below: ")
    for row in range(start,stop): #Loop runs only once.
        
        name = input("Enter name: ")
        gender = input("Enter gender: ")

        try:
            room_no = int(input("Enter room number: "))
        except:
            print("Room number must be an integer value.")
            room_no = int(input("Enter room number: "))
            

        address = input("Enter address: ")
        doctor = input("Enter the name of the doctor handling the case: ")
        department = input("Enter department: ")

        try:
            fee = float(input("Enter the fee to be deposited by the patient (in $) : "))
        except:
            print("Fee should be an integer or a floating point value only.")
            fee = float(input("Enter the fee to be deposited by the patient (in $) : "))
            
        worksheet["B" + str(row)].value = name  
        worksheet["C" + str(row)].value = gender
        worksheet["D" + str(row)].value = room_no
        worksheet["E" + str(row)].value = address
        worksheet["F" + str(row)].value = doctor
        worksheet["G" + str(row)].value = department
        worksheet["H" + str(row)].value = "$"+str(fee)
        print("Records of patient", ID, "have been deleted.")
        workbook.save(r"C:\Users\maashree\Desktop\Admin\patient_database.xlsx")
        
def delete():
    workbook = load_workbook(r"C:\Users\maashree\Desktop\Admin\patient_database.xlsx")
    worksheet = workbook.active

    try: #To avoid non integer input.
        ID = int(input("Enter the patient ID whose records you want to delete: "))
    except:
        print("The ID of patient should be an integer value.")
        ID = int(input("Enter the patient ID whose records you want to delete: "))

    delete = ID + 1
    print("Records of the patient of ID :", worksheet["A"+str(delete)].value, "have been deleted from this database")
    worksheet.delete_rows(delete)
    
    
    worksheet.insert_rows(delete)
    
    workbook.save(r"C:\Users\maashree\Desktop\Admin\patient_database.xlsx")
#Functions to manage doctors using admin mode start from here.
def headingD(): #All functions work similar to the previous set of functions.
    workbook = load_workbook(r"C:\Users\maashree\Desktop\Admin\doctor_database.xlsx")
    worksheet = workbook.active

    worksheet["A1"].value = "Doctor ID"
    worksheet["B1"].value = "Name"
    worksheet["C1"].value = "Specialisation"
    worksheet["D1"].value = "Salary"
    worksheet["E1"].value = "Department"
    worksheet["F1"].value = "No of patients"

    workbook.save(r"C:\Users\maashree\Desktop\Admin\doctor_database.xlsx")
   
def addD():
    workbook = load_workbook(r"C:\Users\maashree\Desktop\Admin\doctor_database.xlsx")
    worksheet = workbook.active
    number = int(input("How many records do you want to add?: "))

    for i in range(number):
        check = lastIDD()
        if check == None:
            ID = 1

        else:
            ID = check+1
            
        name = input("Enter doctor's name : ")
        spec = input("Enter doctor's specialisation : ")
        try:
            sal = float(input("Enter doctor's salary (in $) : "))
        except:
            print("Doctor's salary can only be a floating point number or a whole number.")
            sal = float(input("Enter doctor's salary (in $): "))
        dep = input("Enter doctor's department : ")

        try:
            num = int(input("How many patients is the doctor handling currently? : "))
        except:
            print("Number of patients being handled must be an integer value.")
            num = int(input("How many patients is the doctor handling currently? : "))
        data = [ID, name, spec, "$"+str(sal), dep, num]
        worksheet.append(data)
        workbook.save(r"C:\Users\maashree\Desktop\Admin\doctor_database.xlsx")
        print("The record of doctor", ID, "is added")

def lastIDD():
    workbook = load_workbook(r"C:\Users\maashree\Desktop\Admin\doctor_database.xlsx")
    worksheet = workbook.active
    L = []
    for i in range(2,10000):
        ID = worksheet["A"+str(i)].value
        IDs=str(ID)
        if IDs.isdigit():
            L.append(ID)
            continue
    
    try:
        last = max(L)
        print("The latest ID is : ", last)
        return last

    except:
        print("Database is currently empty, last ID is 0")

        
def displayD():
    workbook = load_workbook(r"C:\Users\maashree\Desktop\Admin\doctor_database.xlsx")
    worksheet = workbook.active

    try:
        ID = int(input("Enter the doctor's ID whose records you want to see : "))
    except:
        print("Doctor's ID should be an integer value.")
        ID = int(input("Enter the doctor's ID whose records you want to see : "))
    header = 1
    start = ID+1
    stop = ID+2
    for row in range(start, stop):
        for col in range(1,7):

            char = get_column_letter(col)
            print(worksheet[char+str(header)].value," : ",worksheet[char+str(row)].value)


def deleteD():
    workbook = load_workbook(r"C:\Users\maashree\Desktop\Admin\doctor_database.xlsx")
    worksheet = workbook.active
    try:
        ID = int(input("Enter the doctor's ID who you want to delete from the database"))
    except:
        print("Doctor's ID should be an integer value.")
        ID = int(input("Enter the doctor's ID who you want to delete from the database"))
    delete = ID+1
    print("Records of the doctor", worksheet["A"+str(delete)].value, "have been deleted from the database")
    worksheet.delete_rows(delete)

    worksheet.insert_rows(delete)
    workbook.save(r"C:\Users\maashree\Desktop\Admin\doctor_database.xlsx")

    
def editD():
    workbook = load_workbook(r"C:\Users\maashree\Desktop\Admin\doctor_database.xlsx")
    worksheet = workbook.active

    try:
        ID = int(input("Enter the ID of the doctor whose records you want to edit : "))
    except:
        print("Doctor's ID should be an integer value.")
        ID = int(input("Enter the ID of the doctor whose records you want to edit : "))
        

    start = ID+1
    stop = ID+2
    for row in range(start,stop):
        
        print("Enter the details to be edited below : ")
        name = input("Enter name of the doctor : ")
        spec = input("Enter doctor's specialisation : ")

        try:
            sal = float(input("Enter doctor's salary (in $): "))
        except:
            print("Doctor's salary must either be a floating point value or an integer.")
            sal = float(input("Enter doctor's salary (in $): "))

        dep = input("Enter doctor's department : ")
        try:
            num = int(input("How many patients is the doctor currently handling? : "))
        except:
            print("No. of pateints being handled must be an integer value.")
            num = int(input("How many patients is the doctor currently handling? : "))
        worksheet["B"+str(row)].value = name
        worksheet["C"+str(row)].value = spec
        worksheet["D"+str(row)].value = "$"+str(sal)
        worksheet["E"+str(row)].value = dep
        worksheet["F"+str(row)].value = num
        print("Records of doctor", ID, "have been edited")
        workbook.save(r"C:\Users\maashree\Desktop\Admin\doctor_database.xlsx")

#Doctor mode functions start from here.
def headingDoc(): #All functions work similar to previous set of functions.
    workbook = load_workbook(r"C:\Users\maashree\Desktop\Doctor\patient_databaseD.xlsx")
    worksheet = workbook.active

    worksheet["A1"].value = "Patient ID"
    worksheet["B1"].value = "Name"
    worksheet["C1"].value = "Gender"
    worksheet["D1"].value = "Room No."
    worksheet["E1"].value = "Age"
    worksheet["F1"].value = "Diagnosis"
    worksheet["G1"].value = "Treatment"
    worksheet["H1"].value = "Duration of Treatment"
    worksheet["I1"].value = "Date of admission"
    worksheet["J1"].value = "Remarks"
    worksheet["K1"].value = "Tests"
    worksheet["L1"].value = "Name of Doctor"
    workbook.save(r"C:\Users\maashree\Desktop\Doctor\patient_databaseD.xlsx")


def addDoc():
    workbook = load_workbook(r"C:\Users\maashree\Desktop\Doctor\patient_databaseD.xlsx")
    worksheet = workbook.active
    number = int(input("How many records do you want to add? : "))
    
    for i in range(number):
        check = lastIDDoc()

        if check == None:
            ID = 1

        else:
            ID = check+1

        name = input("Enter name : ")
        gender = input("Enter gender : ")

        try:
            room_no =  int(input("Enter room no. : "))
        except:
            print("Room no. should be an integer value.")
            room_no =  int(input("Enter room no. : "))


        try:
            age = int(input("Enter age : "))
        except:
            print("Age should be an integer value.")

        diagnosis = input("Enter the diagnosis : ")
        treatment = input("Enter treatment : ")

        try:
            dur = float(input("Enter duration of treatment (in months) : "))
        except:
            print("Duration of treatment should either be an integer or a floating point value.")
            dur = float(input("Enter duration of treatment (in months) : "))

        date = input("Enter date (dd/mm/yy) : ")
        remarks = input("Enter remarks : ")
        tests = input("Enter tests to be done on patient : ")
        doctor = input("Enter the name of doctor handling this patient : ")
        
        data = [ID, name, gender, room_no, age, diagnosis, treatment, str(dur)+"Month(s)", date, remarks, tests, doctor]
        worksheet.append(data)
        workbook.save(r"C:\Users\maashree\Desktop\Doctor\patient_databaseD.xlsx")
        print("Record of patient", ID, "is added")


def lastIDDoc():
    workbook = load_workbook(r"C:\Users\maashree\Desktop\Doctor\patient_databaseD.xlsx")
    worksheet = workbook.active

    L = []
    for i in range(2,10000):
        ID = worksheet["A"+str(i)].value
        IDs = str(ID)
        if IDs.isdigit():
            L.append(ID)
            continue

    try:
        last = max(L)
        print("Last ID is : ", last)
        return last
    except:
        print("Database is currently empty, last ID is 0")


def displayDoc():
    workbook = load_workbook(r"C:\Users\maashree\Desktop\Doctor\patient_databaseD.xlsx")
    worksheet = workbook.active

    try:
        ID = int(input("Enter the patient's ID whose records you want to see : "))
    except:
        print("Patient ID should be an integer value.")
        ID = int(input("Enter the patient's ID whose records you want to see : "))
    start = ID+1
    stop = ID+2
    header = 1
    for row in range(start, stop):
        for col in range(1, 13):
            char = get_column_letter(col)
            print(worksheet[char+str(header)].value, " : ", worksheet[char+str(row)].value)


def deleteDoc():
    workbook = load_workbook(r"C:\Users\maashree\Desktop\Doctor\patient_databaseD.xlsx")
    worksheet = workbook.active
    try:
        ID = int(input("Enter the ID of patient whose records you want to delete : "))
    except:
        print("Patient ID should be an integer value.")
        ID = int(input("Enter the ID of patient whose records you want to delete : "))
    delete = ID+1
    print("Records of patient", worksheet["A"+str(delete)].value, "have been deleted from the database")
    worksheet.delete_rows(delete)

    worksheet.insert_rows(delete)
    workbook.save(r"C:\Users\maashree\Desktop\Doctor\patient_databaseD.xlsx")


def editDoc():
    workbook = load_workbook(r"C:\Users\maashree\Desktop\Doctor\patient_databaseD.xlsx")
    worksheet = workbook.active
    try:
        ID = int(input("Enter the ID of the patient whose records you want to edit : "))
    except:
        print("Patient's ID should be an integer value.")
        ID = int(input("Enter the ID of the patient whose records you want to edit : "))
        
    start = ID+1
    stop = ID+2
    for row in range(start,stop):

        print("Enter the details to be edited below : ")
        name = input("Enter name : ")
        gender = input("Enter gender : ")
        try:
            room_no =  int(input("Enter room no. : "))
        except:
            print("Room no. should be an integer value.")
            room_no =  int(input("Enter room no. : "))
            
        try:
            age = int(input("Enter age : "))
        except:
            print("Age should be an integer value.")
            age = int(input("Enter age : "))
            
        diagnosis = input("Enter the diagnosis : ")
        treatment = input("Enter treatment : ")
        try:
            dur = float(input("Enter duration of treatment (in months) : "))
        except:
            print("Duration of treatment should either be an integer or a floating point value.")
            dur = float(input("Enter duration of treatment (in months) : "))
        date = input("Enter date (dd/mm/yy) : ")
        remarks = input("Enter remarks : ")
        tests = input("Enter tests to be done on patient : ")
        doctor = input("Enter the name of doctor handling this patient : ")

        worksheet["B"+str(row)].value = name
        worksheet["C"+str(row)].value = gender
        worksheet["D"+str(row)].value = room_no
        worksheet["E"+str(row)].value = age
        worksheet["F"+str(row)].value = diagnosis
        worksheet["G"+str(row)].value = treatment
        worksheet["H"+str(row)].value = str(dur)+"months"
        worksheet["I"+str(row)].value = date
        worksheet["J"+str(row)].value = remarks
        worksheet["K"+str(row)].value = tests
        worksheet["L"+str(row)].value = doctor

        workbook.save(r"C:\Users\maashree\Desktop\Doctor\patient_databaseD.xlsx")
        print("Records of patient", ID, "have been edited")
        

print("-------------------------------------------------------------------------")
print("|Enter 1 for Admin Mode|\n|Enter 2 for Doctor Mode|")
print("-------------------------------------------------------------------------")

admin_doctor_mode = input("Enter your mode:") #Decides between admin mode and doctor mode.

if admin_doctor_mode == "1": #All code inside this block is for admin mode
    print("PLEASE DO NOT PERFORM ANY WRITING ACTION ON DATA DIRECTLY THROUGH EXCEL")
    print("*********************************************************************************")
    password = input("Enter your password : ")

    while True:

        if password == "1234": #Executes only if password entered is 1234.
            print("--------------------------------------------------------------------------------------")
            print("|To manage patients enter 1|\n|To manage doctors enter 2|")
            print("---------------------------------------------------------------------------------------")
            adminchoice = input("Enter your choice : ") #Decides if admin wants to manage patients or doctors.
            

            if adminchoice == "1": #All code inside this block is for admin to manage patients.

                heading()

                print("-----------------------------------------------------------------------------------------------------")
                print("To add a new patient enter 1")
                print("To display a patient enter 2")
                print("To delete a patient enter 3")
                print("To edit a patient's data enter 4")
                print("To go back enter E")

                admin_action = input("Enter your choice : ")
                admin_action = admin_action.upper() #To handle if user enters a lowercase e
                if admin_action == "1":
                    add()
                    
                elif admin_action == "2":
                    display()
                    
                elif admin_action == "3":
                    delete()

                elif admin_action == "4":
                    edit()

                elif admin_action == "E":
                    break

                else:
                    print("Please enter a valid choice")
                    
            if adminchoice == "2": #All code in this block is for admin  to manage doctors.
                headingD()

                print("-----------------------------------------------------------------------------------------------------")
                print("To add a new doctor enter 1")
                print("To display a doctor enter 2")
                print("To delete a doctor enter 3")
                print("To edit a doctor's data enter 4")
                print("To go back enter E")

                admin_action = input("Enter your choice : ")
                admin_action = admin_action.upper()

                if admin_action == "1":
                    addD()

                elif admin_action == "2":
                    displayD()

                elif admin_action == "3":
                    deleteD()

                elif admin_action == "4":
                    editD()

                elif admin_action == "E":
                    break
                    
                else:
                    print("Enter a valid choice")
        else:
            print("ACCESS DENIED. INCORRECT PASSWORD.")
            break        
        
#Admin mode completed here.            
if admin_doctor_mode == "2": #All code in this block is for doctor mode.
    print("PLEASE DO NOT PERFORM ANY WRITING ACTION ON DATA DIRECTLY THROUGH EXCEL")
    print("*******************************************************************************************************************")
    password = input("Enter your password : ")
    headingDoc()
    while True:
        if password == "5678": #Excecutes only if password entered is 5678.
            print("To add a patient enter 1 : ")
            print("To display a patient enter 2 : ")
            print("To delete a patient's data enter 3 : ")
            print("To edit a patient's data enter 4 : ")
            print("To go back enter E : ")

            doctor_action = input("Enter your choice : ")
            doctor_action = doctor_action.upper()

            if doctor_action == "1":
                addDoc()

            elif doctor_action == "2":
                displayDoc()

            elif doctor_action == "3":
                deleteDoc()
                
            elif doctor_action == "4":
                editDoc()

            elif doctor_action == "E":
                break

            else:
                print("Enter a valid choice")
        else:
            print("ACCESS DENIED. INCORRECT PASSWORD.")
            break
